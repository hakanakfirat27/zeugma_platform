# reports/services/company_import.py
"""
Company Import Service - Multi-Sheet Excel Import

Imports companies from a single Excel file with 10 sheets (one per category).
STRICT MERGE: Only merges companies if ALL 29 company fields match exactly.
Generates report of potential duplicates for manual review.
"""

import re
import hashlib
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db import transaction, OperationalError
from django.utils import timezone
from reports.company_models import (
    Company, ProductionSite, ProductionSiteVersion,
    CompanyStatus, CompanyHistory
)


# =============================================================================
# RETRY HELPER FOR SQLITE LOCKING
# =============================================================================

def retry_on_locked(func, max_retries=5, base_delay=0.1):
    """
    Retry a function if database is locked (SQLite limitation).
    Uses exponential backoff.
    """
    for attempt in range(max_retries):
        try:
            return func()
        except OperationalError as e:
            if 'database is locked' in str(e) and attempt < max_retries - 1:
                delay = base_delay * (2 ** attempt)  # Exponential backoff
                time.sleep(delay)
            else:
                raise
    return None


# =============================================================================
# FIELD DEFINITIONS (from fields.py)
# =============================================================================

# Company-level fields (stored in Company model) - EXACTLY 29 fields
COMMON_FIELDS = ['company_name', 'address_1', 'address_2', 'address_3', 'address_4', 
                 'region', 'country', 'geographical_coverage', 'phone_number', 
                 'company_email', 'website', 'accreditation', 'parent_company']

CONTACT_FIELDS = ['title_1', 'initials_1', 'surname_1', 'position_1', 
                  'title_2', 'initials_2', 'surname_2', 'position_2', 
                  'title_3', 'initials_3', 'surname_3', 'position_3', 
                  'title_4', 'initials_4', 'surname_4', 'position_4']

COMPANY_FIELDS = set(COMMON_FIELDS + CONTACT_FIELDS)

# Category-specific version fields
INJECTION_FIELDS = ['custom', 'proprietary_products', 'in_house', 'ps', 'san', 'abs', 'ldpe', 'lldpe', 'hdpe', 'pp', 'pom', 'pa', 'pa12', 'pa11', 'pa66', 'pa6', 'pmma', 'pc', 'ppo', 'peek', 'pet', 'pbt', 'psu', 'pvc', 'tpes', 'tpes_type', 'thermosets', 'bioresins', 'other_materials', 'main_materials', 'polymer_range_number', 'polymer_range', 'automotive', 'automotive_under_the_bonnet', 'automotive_interior', 'automotive_exterior', 'automotive_description', 'electrical', 'electrical_components', 'appliances', 'telectronics', 'electrical_description', 'houseware', 'housewares_non_electrical', 'pet_accessories', 'houseware_description', 'medical', 'medical_devices', 'medical_containers', 'medical_description', 'packaging', 'crates_boxes', 'preform', 'pails', 'thin_wall_food_packaging', 'cosmetics_packaging', 'caps_closures', 'lids', 'cartons_caps', 'aerosol_overcaps', 'pumps', 'push_on_caps', 'screwcaps', 'packaging_description', 'building', 'horticulture_agriculture', 'sport_leisure', 'toys', 'dvd_cds', 'pipe_fittings', 'personal_care', 'furniture', 'other_products', 'main_applications', 'clean_room', 'tool_design', 'tool_manufacture', 'pad_printing', 'hot_foil_stamping', 'insert_moulding', 'painting', 'inmould_labelling', 'electroplating_metalizing', 'twin_multi_shot_moulding', 'gas_water_assisted_moulding', 'three_d_printing', 'other_services', 'minimal_lock_tonnes', 'maximum_lock_tonnes', 'minimum_shot_grammes', 'maximum_shot_grammes', 'number_of_machines', 'machinery_brand']

BLOW_FIELDS = ['in_house', 'custom', 'proprietary_products', 'hdpe', 'ldpe', 'lldpe', 'pet', 'apet', 'petg', 'cpet', 'pp', 'pvc', 'pc', 'other_materials', 'main_materials', 'polymer_range_number', 'polymer_range', 'under_1_litre', 'from_1_to_5_litres', 'from_5_to_25_litres', 'from_25_to_220_litres', 'over_220_litres', 'auto_fuel_tanks', 'automotive', 'cosmetics_packaging', 'food_drink', 'household_chemicals', 'industrial_chemicals', 'medical', 'technical_moulding', 'toys', 'tubes', 'other_products', 'main_applications', 'multilayer', 'clean_room', 'tool_design', 'tool_manufacture', 'product_development', 'just_in_time', 'assembly', 'filling', 'labelling', 'welding', 'embossing', 'pad_printing', 'silk_screen_printing', 'offset_printing', 'other_printing', 'printing', 'number_of_colours', 'design', 'other_services', 'extrusion_blow_moulding_machines', 'injection_blow_moulding_machines', 'injection_stretch_blow_moulding_stage_1_machines', 'injection_stretch_blow_moulding_stage_2_machines', 'buy_in_preform', 'buy_in_preform_percentage', 'number_of_machines', 'machinery_brand']

ROTO_FIELDS = ['custom', 'in_house', 'proprietary_products', 'ldpe', 'lldpe', 'xlpe', 'hdpe', 'eva', 'pvc', 'pa', 'pc', 'lldpe_c4', 'lldpe_c6', 'lldpe_c8', 'other_materials', 'main_materials', 'polymer_range_number', 'polymer_range', 'horticulture_agriculture', 'marine_fishing', 'toys', 'sport_leisure', 'road_furniture', 'footballs', 'automotive_vehicles_trucks', 'ibcs', 'boats_canoes_kayaks', 'heating_oil_diesel_tanks', 'chemical_un_tanks', 'septic_tanks', 'water_tank_other_tanks', 'waste_container_bottle_banks', 'underground_fittings', 'pos_mannequins_displays', 'materials_handling_boxes', 'materials_handling_pallets', 'other_products', 'main_applications', 'printing', 'tool_design', 'tool_manufacture', 'product_development', 'design', 'other_services', 'minimum_size', 'maximum_size', 'number_of_machines']

PE_FILM_FIELDS = ['ldpe', 'lldpe', 'lldpe_c4', 'lldpe_c6', 'lldpe_c8', 'mlldpe', 'hdpe', 'eva', 'eaa_eba', 'ionomer', 'pa', 'pp', 'pvc', 'bioresins', 'other_materials', 'main_materials', 'polymer_range_number', 'polymer_range', 'minimum_width_mm', 'maximum_width_mm', 'multilayer', 'number_of_layers', 'cast_lines', 'blown_lines', 'machinery_brand', 'film_on_reel', 'agricultural_film', 'agricultural_silage', 'agricultural_mulch', 'agricultural_greenhouse', 'shrink_fill_collation', 'shrink_film_pallet', 'stretch_film', 'stretch_hood', 'building_film', 'hygiene_film', 'laminating_film', 'freezer_film', 'other_films', 'carrier_bags', 'biodegradable_bags', 'other_bags', 'industrial_sacks', 'refuse_sacks', 'other_sacks', 'other_products', 'main_applications', 'product_lamination', 'printing', 'flexo_printing', 'gravure_printing', 'number_of_colours', 'other_services']

SHEET_FIELDS = ['custom', 'in_house', 'proprietary_products', 'ps', 'abs', 'pp', 'ldpe', 'lldpe', 'hdpe', 'pmma', 'pc', 'pvc', 'pet', 'petg', 'uhmwpe', 'recycled_materials', 'pla', 'other_bioresins', 'other_materials', 'main_materials', 'polymer_range_number', 'polymer_range', 'compound_in_house', 'buy_in_compounds', 'minimum_gauge_mm', 'maximum_gauge_mm', 'minimum_width_mm', 'maximum_width_mm', 'flame_retardant_sheet', 'embossed_sheet', 'coated_sheet', 'laminated_sheet', 'foamed_sheet', 'corrugated_sheet', 'profiled_sheet', 'coextruded_sheet', 'number_of_layers', 'plain_sheet', 'cross_linked_sheet', 'other_type_of_sheet', 'extrusion_process', 'number_of_extrusion_lines', 'coextrusion_process', 'number_of_coextrusion_lines', 'calendering_process', 'number_of_calendering_lines', 'pressing_process', 'number_of_pressed_lines', 'liquid_cell_casting', 'number_of_lcc_line', 'in_line_process', 'glazing', 'lighting', 'membrane', 'insulation', 'flooring', 'construction_other', 'display', 'pannelling', 'automotive', 'aerospace', 'appliances', 'sport_leisure', 'stationery_supplies', 'other_industrial_markets', 'chilled_food', 'fast_food', 'meat_fish', 'yellow_fats', 'frozen_food', 'bakery_confectionary', 'ovenable', 'fruit_vegetable', 'ambient_food', 'cups', 'retail_pos_packaging', 'medical', 'protective_packaging', 'other_food_packaging', 'non_food_packaging', 'other_products', 'main_applications']

PIPE_FIELDS = ['below_ground_pressure', 'below_ground_non_pressure', 'above_ground_pressure', 'above_ground_non_pressure', 'other_pipes', 'water_supply_distribution', 'gas_transmission_distribution', 'drainage_sewerage', 'storm_water_drainage', 'road_railways_drainage', 'agricultural_drainage', 'cable_protection', 'internal_soil_waste_sewerage', 'internal_hot_cold_plumbing_heating', 'roof_gutter_systems', 'irrigation', 'air_conditioning', 'other_products', 'main_applications', 'up_to_32_mm', 'between_33_63_mm', 'between_64_90_mm', 'between_91_109_mm', 'between_110_160_mm', 'between_161_250_mm', 'between_251_400_mm', 'between_401_500_mm', 'between_501_630_mm', 'between_631_1000_mm', 'between_1001_1200_mm', 'over_1200_mm', 'lldpe', 'ldpe', 'hdpe', 'mdpe', 'xlpe', 'pe80', 'pe100', 'pert', 'pvc', 'pvcc', 'polybutylene', 'pp', 'abs', 'pa', 'modified_pvc', 'pp_r', 'oriented_pvc', 'pe100rc', 'other_materials', 'main_materials', 'polymer_range_number', 'polymer_range', 'coextruded_foam_core', 'twin_wall', 'multilayer_polymeric', 'multilayer_polymer_metal', 'reinforced_wall', 'solid_wall', 'other_technologies', 'number_of_machines']

TUBE_HOSE_FIELDS = ['custom', 'in_house', 'proprietary_products', 'pvc', 'rigid_pvc', 'flexible_pvc', 'ldpe', 'hdpe', 'lldpe', 'pp', 'ps', 'pom', 'pa', 'abs', 'tpes', 'tpe_e', 'sebs', 'tpu', 'tpv', 'pmma', 'pc', 'pbt', 'other_materials', 'main_materials', 'polymer_range_number', 'polymer_range', 'compound_in_house', 'buy_in_compounds', 'automotive', 'medical', 'food_drink', 'horticulture_agriculture', 'electrical_conduit', 'other_products', 'main_applications', 'hydraulic', 'pneumatic', 'flexible', 'rigid_shapes', 'monolayer', 'multilayer', 'extrusion_foam', 'corrugated_extrusion', 'twin_wall', 'bi_colour', 'other_technologies', 'machinery_brand', 'number_of_machines', 'minimum_diameter_mm', 'maximum_diameter_mm', 'clean_room', 'tool_design', 'just_in_time', 'high_frequency_welding', 'ultrasonic_welding', 'other_welding', 'pad_printing', 'silk_screen_printing', 'hot_foil_stamping', 'other_printing', 'assembly', 'machining', 'recycling', 'other_services']

PROFILE_FIELDS = ['custom', 'in_house', 'proprietary_products', 'pvc', 'rigid_pvc', 'flexible_pvc', 'ldpe', 'hdpe', 'lldpe', 'pp', 'ps', 'pom', 'pa', 'abs', 'tpes', 'tpe_e', 'sebs', 'tpu', 'tpv', 'pmma', 'pc', 'pbt', 'wpc', 'other_materials', 'main_materials', 'compound_in_house', 'buy_in_compounds', 'polymer_range_number', 'polymer_range', 'door', 'window', 'seals_gaskets', 'soffits_bargeboard_compact', 'soffits_bargeboard_foamed', 'cladding', 'furniture', 'lighting', 'blinds', 'trunking', 'floor_panels', 'decking', 'other_products', 'main_applications', 'monolayer', 'multilayer', 'extrusion_foam', 'corrugated_extrusion', 'twin_wall', 'bi_colour', 'other_technologies', 'machinery_brand', 'number_of_machines', 'minimum_diameter_mm', 'maximum_diameter_mm', 'clean_room', 'tool_design', 'just_in_time', 'high_frequency_welding', 'ultrasonic_welding', 'other_welding', 'pad_printing', 'silk_screen_printing', 'hot_foil_stamping', 'other_printing', 'assembly', 'machining', 'recycling', 'other_services', 'shutters']

CABLE_FIELDS = ['pvc', 'ldpe', 'cellular_pe', 'hdpe_mdpe', 'xlpe', 'pp', 'tpes', 'elastomers', 'lldpe', 'lsf0h', 'pbt', 'silicone', 'fluoropolymers', 'cpe', 'cspe', 'epr', 'epdm', 'eva', 'other_materials', 'main_materials', 'polymer_range_number', 'polymer_range', 'low_voltage_1k_insulation', 'medium_voltage_1_36k_insulation', 'high_voltage_36k_insulation', 'metallic_insulation', 'optical_insulation', 'automotive_insulation', 'mining_insulation', 'data_cable_insulation', 'computer_cable_insulation', 'microphone_cable_insulation', 'plenum_insulation', 'low_voltage_1k_jacketing', 'medium_voltage_1_36k_jacketing', 'high_voltage_36k_jacketing', 'metallic_jacketing', 'optical_jacketing', 'automotive_jacketing', 'mining_jacketing', 'microphone_cable_jacketing', 'computer_cable_jacketing', 'data_cable_jacketing', 'plenum_jacketing', 'other_products', 'main_applications', 'number_of_machines']

COMPOUNDER_FIELDS = ['lldpe', 'ldpe', 'hdpe', 'ps', 'pp', 'rigid_pvc', 'flexible_pvc', 'pa', 'pc', 'abs', 'san', 'pet', 'acetal', 'pbt', 'tpes', 'other_materials', 'main_materials', 'reprocessing', 'colour_compounds', 'flame_retardant_compounds', 'mineral_filled_compounds', 'glass_filled_compounds', 'elastomer_modified_compounds', 'cross_linked_compounds', 'carbon_fibre_compounds', 'natural_fibre_compounds', 'other_compounds', 'compounds_percentage', 'compounds_applications', 'black_masterbatch', 'white_masterbatch', 'colour_masterbatch', 'additive_masterbatch', 'liquid_masterbatch', 'other_masterbatches', 'masterbatch_percentage', 'masterbatches_applications', 'main_applications', 'number_of_machines', 'twin_screw_extruders', 'single_screw_extruders', 'batch_mixers', 'polymer_producer', 'production_volume_number', 'polymer_range']

# Map category to its fields
CATEGORY_FIELDS = {
    'INJECTION': INJECTION_FIELDS,
    'BLOW': BLOW_FIELDS,
    'ROTO': ROTO_FIELDS,
    'PE_FILM': PE_FILM_FIELDS,
    'SHEET': SHEET_FIELDS,
    'PIPE': PIPE_FIELDS,
    'TUBE_HOSE': TUBE_HOSE_FIELDS,
    'PROFILE': PROFILE_FIELDS,
    'CABLE': CABLE_FIELDS,
    'COMPOUNDER': COMPOUNDER_FIELDS,
}

# All version fields (union of all categories)
ALL_VERSION_FIELDS = set()
for fields in CATEGORY_FIELDS.values():
    ALL_VERSION_FIELDS.update(fields)


# =============================================================================
# SHEET NAME TO CATEGORY MAPPING
# =============================================================================

SHEET_TO_CATEGORY = {
    'injection moulders': 'INJECTION',
    'injection': 'INJECTION',
    'blow moulders': 'BLOW',
    'blow': 'BLOW',
    'roto moulders': 'ROTO',
    'roto': 'ROTO',
    'rotational moulders': 'ROTO',
    'pe film extruders': 'PE_FILM',
    'pe film': 'PE_FILM',
    'film extruders': 'PE_FILM',
    'sheet extruders': 'SHEET',
    'sheet': 'SHEET',
    'pipe extruders': 'PIPE',
    'pipe': 'PIPE',
    'tube & hose extruders': 'TUBE_HOSE',
    'tube and hose extruders': 'TUBE_HOSE',
    'tube & hose': 'TUBE_HOSE',
    'tube': 'TUBE_HOSE',
    'profile extruders': 'PROFILE',
    'profile': 'PROFILE',
    'cable extruders': 'CABLE',
    'cable': 'CABLE',
    'compounders': 'COMPOUNDER',
    'compounder': 'COMPOUNDER',
}


# =============================================================================
# HEADER NORMALIZATION MAPPING
# =============================================================================

def normalize_header(header):
    """Normalize Excel header to field name."""
    if not header:
        return None
    
    # Lowercase and strip
    header = str(header).lower().strip()
    
    # Replace special chars with underscores
    header = re.sub(r'[^a-z0-9]+', '_', header)
    header = header.strip('_')
    
    # Common transformations
    mappings = {
        'company_name': 'company_name',
        'company': 'company_name',
        'address_1': 'address_1',
        'address1': 'address_1',
        'address_2': 'address_2',
        'address2': 'address_2',
        'address_3': 'address_3',
        'address3': 'address_3',
        'address_4': 'address_4',
        'address4': 'address_4',
        'phone_number': 'phone_number',
        'phone': 'phone_number',
        'tel': 'phone_number',
        'telephone': 'phone_number',
        'company_email': 'company_email',
        'email': 'company_email',
        'e_mail': 'company_email',
        'title_1': 'title_1',
        'title1': 'title_1',
        'initials_1': 'initials_1',
        'initials1': 'initials_1',
        'surname_1': 'surname_1',
        'surname1': 'surname_1',
        'position_1': 'position_1',
        'position1': 'position_1',
        'title_2': 'title_2',
        'title2': 'title_2',
        'initials_2': 'initials_2',
        'initials2': 'initials_2',
        'surname_2': 'surname_2',
        'surname2': 'surname_2',
        'position_2': 'position_2',
        'position2': 'position_2',
        'title_3': 'title_3',
        'title3': 'title_3',
        'initials_3': 'initials_3',
        'initials3': 'initials_3',
        'surname_3': 'surname_3',
        'surname3': 'surname_3',
        'position_3': 'position_3',
        'position3': 'position_3',
        'title_4': 'title_4',
        'title4': 'title_4',
        'initials_4': 'initials_4',
        'initials4': 'initials_4',
        'surname_4': 'surname_4',
        'surname4': 'surname_4',
        'position_4': 'position_4',
        'position4': 'position_4',
        'geographical_coverage': 'geographical_coverage',
        'geo_coverage': 'geographical_coverage',
    }
    
    # Return mapped value if exists
    if header in mappings:
        return mappings[header]
    
    # Otherwise return header as-is (it should match field name already)
    return header


# =============================================================================
# FIELD TYPE DETECTION
# =============================================================================

# Integer and Float fields
NUMERIC_FIELDS = {
    'minimal_lock_tonnes', 'maximum_lock_tonnes',
    'minimum_shot_grammes', 'maximum_shot_grammes',
    'number_of_machines',
    'extrusion_blow_moulding_machines', 'injection_blow_moulding_machines',
    'injection_stretch_blow_moulding_stage_1_machines', 'injection_stretch_blow_moulding_stage_2_machines',
    'buy_in_preform_percentage', 'number_of_colours',
    'minimum_size', 'maximum_size',
    'minimum_width_mm', 'maximum_width_mm',
    'minimum_gauge_mm', 'maximum_gauge_mm',
    'minimum_diameter_mm', 'maximum_diameter_mm',
    'number_of_layers', 'cast_lines', 'blown_lines',
    'number_of_extrusion_lines', 'number_of_coextrusion_lines',
    'number_of_calendering_lines', 'number_of_pressed_lines', 'number_of_lcc_line',
    'twin_screw_extruders', 'single_screw_extruders', 'batch_mixers',
    'production_volume_number', 'polymer_range_number',
    'compounds_percentage', 'masterbatch_percentage',
    # Additional numeric fields from model
    'number_of_recycling_lines', 'single_screws', 'twin_screws',
}

# For backwards compatibility
INTEGER_FIELDS = NUMERIC_FIELDS

# Text fields (not boolean, not integer)
TEXT_FIELDS = {
    'other_materials', 'main_materials', 'polymer_range',
    'automotive_description', 'electrical_description', 'houseware_description',
    'medical_description', 'packaging_description',
    'other_products', 'main_applications', 'other_services',
    'tpes_type', 'machinery_brand',
    'other_films',  # This IS a text field
    'other_type_of_sheet', 'other_technologies',
    'other_pipes', 'other_printing',
    'compounds_applications', 'masterbatches_applications',
    'other_compounds', 'other_masterbatches',
    'other_bioresins', 'other_food_packaging', 'non_food_packaging',
    'construction_other', 'other_industrial_markets',
    'other_welding', 'in_line_process', 'minimum_size', 'maximum_size',
    # Recycler text fields
    'other_contaminations', 'other_packaging', 'other_building',
    'other_automotive', 'other_weee', 'other_markets', 'main_source',
    'other_mechanical_process', 'finished_products', 'volume_of_recycled_product',
    'minimum_volume_waste_accepted', 'maximum_capacity',
}

# NOTE: other_bags and other_sacks are BOOLEAN fields, not text!
# They were incorrectly listed here before.


def is_boolean_field(field_name):
    """Check if a field is a boolean field."""
    return field_name not in INTEGER_FIELDS and field_name not in TEXT_FIELDS


def parse_boolean(value):
    """Parse a value as boolean. PostgreSQL is strict, so we must return True/False only."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        value = value.strip().lower()
        if value in ('yes', 'y', 'true', '1', 'x', '✓', '✔', 'checked'):
            return True
        elif value in ('no', 'n', 'false', '0', '', 'none', 'null'):
            return False
        # Default to False for any other string
        return False
    return False


def parse_integer(value):
    """Parse a value as integer."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value) if value else None
    if isinstance(value, str):
        value = value.strip()
        # Handle Y/N which might be mistakenly in numeric columns
        if value.lower() in ('y', 'yes', 'n', 'no', 'true', 'false', ''):
            return None
        value = re.sub(r'[^\d-]', '', value)
        if value:
            try:
                return int(value)
            except ValueError:
                return None
    return None


def clean_version_data_for_postgres(data, model_class):
    """
    Clean version data to ensure PostgreSQL compatibility.
    Ensures all boolean fields are True/False, not strings.
    """
    from django.db import models
    
    cleaned = {}
    
    # Get field types from the model
    field_types = {}
    for field in model_class._meta.get_fields():
        if hasattr(field, 'get_internal_type'):
            field_types[field.name] = field.get_internal_type()
    
    for key, value in data.items():
        field_type = field_types.get(key)
        
        if field_type == 'BooleanField':
            # Force conversion to Python bool
            cleaned[key] = parse_boolean(value)
        elif field_type in ('IntegerField', 'PositiveIntegerField'):
            # Force conversion to int or None
            cleaned[key] = parse_integer(value)
        elif field_type in ('FloatField', 'DecimalField'):
            # Force conversion to float or None
            if value is None or (isinstance(value, str) and value.strip() == ''):
                cleaned[key] = None
            elif isinstance(value, str) and value.strip().lower() in ('y', 'yes', 'n', 'no'):
                cleaned[key] = None
            else:
                try:
                    cleaned[key] = float(value) if value else None
                except (ValueError, TypeError):
                    cleaned[key] = None
        else:
            # Keep as-is for other field types
            cleaned[key] = value
    
    return cleaned


def normalize_value(value):
    """Normalize a value for comparison (lowercase, strip whitespace)."""
    if value is None:
        return ''
    return str(value).lower().strip()


def get_company_hash(row_data):
    """
    Generate a hash of all 29 company fields for exact matching.
    This ensures we only merge if ALL fields match exactly.
    """
    all_fields = COMMON_FIELDS + CONTACT_FIELDS
    values = []
    for field in all_fields:
        value = normalize_value(row_data.get(field, ''))
        values.append(value)
    
    # Create hash of all values
    combined = '|'.join(values)
    return hashlib.md5(combined.encode('utf-8')).hexdigest()


def get_simple_key(row_data):
    """
    Get simple key (name + address_1 + country) for potential duplicate detection.
    """
    name = normalize_value(row_data.get('company_name', ''))
    address = normalize_value(row_data.get('address_1', ''))
    country = normalize_value(row_data.get('country', ''))
    return (name, address, country)


def find_differences(row_data1, row_data2):
    """Find which of the 29 company fields differ between two records."""
    all_fields = COMMON_FIELDS + CONTACT_FIELDS
    differences = []
    for field in all_fields:
        val1 = normalize_value(row_data1.get(field, ''))
        val2 = normalize_value(row_data2.get(field, ''))
        if val1 != val2:
            differences.append({
                'field': field,
                'value1': row_data1.get(field, ''),
                'value2': row_data2.get(field, '')
            })
    return differences


# =============================================================================
# IMPORT SERVICE
# =============================================================================

class CompanyImportService:
    """
    Service to import companies from Excel file with multiple sheets.
    Each sheet represents a category (Injection, Blow, etc.).
    
    STRICT MERGE: Only merges companies if ALL 29 fields match exactly.
    Generates report of potential duplicates (same name+address+country but different fields).
    """
    
    @classmethod
    def import_from_excel(cls, file_path, user, report_path=None):
        """
        Import companies from Excel file with multiple sheets.
        
        Args:
            file_path: Path to Excel file
            user: User performing the import
            report_path: Optional path to save potential duplicates report
        """
        results = {
            'companies_created': 0,
            'companies_updated': 0,
            'sites_created': 0,
            'versions_created': 0,
            'merged_count': 0,  # Exact matches (all 29 fields)
            'potential_duplicates': [],  # Same name+address+country but different fields
            'errors': [],
            'sheets_processed': [],
            'total_rows': 0,
        }
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            results['errors'].append({'error': f'Failed to open Excel file: {str(e)}'})
            return results
        
        # Cache 1: Exact match by hash of all 29 fields
        # Key: hash of all 29 fields, Value: (Company, row_data)
        exact_match_cache = {}
        
        # Cache 2: Simple key for potential duplicate detection
        # Key: (name, address_1, country), Value: list of (Company, row_data, category)
        simple_key_cache = {}
        
        # PRE-LOAD existing companies into cache
        print("Loading existing companies into cache...")
        existing_companies = Company.objects.all()
        for comp in existing_companies:
            # Build row_data from company
            row_data = {}
            for field in COMMON_FIELDS + CONTACT_FIELDS:
                row_data[field] = getattr(comp, field, '') or ''
            
            # Add to exact match cache
            comp_hash = get_company_hash(row_data)
            exact_match_cache[comp_hash] = (comp, row_data)
            
            # Add to simple key cache
            simple_key = get_simple_key(row_data)
            if simple_key not in simple_key_cache:
                simple_key_cache[simple_key] = []
            simple_key_cache[simple_key].append((comp, row_data, None))
        
        print(f"Loaded {len(existing_companies)} existing companies into cache")
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Determine category from sheet name
            category = cls._get_category_from_sheet_name(sheet_name)
            if not category:
                results['errors'].append({
                    'sheet': sheet_name,
                    'error': f'Unknown category for sheet: {sheet_name}'
                })
                continue
            
            results['sheets_processed'].append({
                'name': sheet_name,
                'category': category
            })
            
            # Get valid fields for this category
            valid_version_fields = set(CATEGORY_FIELDS.get(category, []))
            
            # Get headers from first row
            headers = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value:
                    field_name = normalize_header(cell_value)
                    headers.append((col, field_name))
            
            # Process each data row
            for row_idx in range(2, sheet.max_row + 1):
                results['total_rows'] += 1
                
                try:
                    # Extract row data
                    row_data = {}
                    for col, field_name in headers:
                        if field_name:
                            cell_value = sheet.cell(row=row_idx, column=col).value
                            row_data[field_name] = cell_value
                    
                    # Skip rows without company name
                    company_name = row_data.get('company_name')
                    if not company_name or str(company_name).strip() == '':
                        continue
                    
                    # Normalize company data
                    company_data = cls._extract_company_data(row_data)
                    
                    # Get hashes/keys
                    comp_hash = get_company_hash(company_data)
                    simple_key = get_simple_key(company_data)
                    
                    # Check for EXACT match (all 29 fields)
                    if comp_hash in exact_match_cache:
                        company, _ = exact_match_cache[comp_hash]
                        results['merged_count'] += 1
                    else:
                        # Check for potential duplicate (same name+address+country but different fields)
                        is_potential_duplicate = False
                        if simple_key in simple_key_cache:
                            for existing_comp, existing_data, existing_cat in simple_key_cache[simple_key]:
                                differences = find_differences(company_data, existing_data)
                                if differences:
                                    # This is a potential duplicate - record it
                                    is_potential_duplicate = True
                                    results['potential_duplicates'].append({
                                        'new_record': {
                                            'category': category,
                                            'sheet': sheet_name,
                                            'row': row_idx,
                                            'data': {f: company_data.get(f, '') for f in COMMON_FIELDS + CONTACT_FIELDS}
                                        },
                                        'existing_record': {
                                            'company_id': str(existing_comp.company_id),
                                            'unique_key': existing_comp.unique_key,
                                            'categories': list(existing_comp.production_sites.values_list('category', flat=True)),
                                            'data': existing_data
                                        },
                                        'differences': differences
                                    })
                                    break  # Only record first match
                        
                        # Create new company (whether potential duplicate or not)
                        company_data['company_name_normalized'] = normalize_value(company_data.get('company_name', ''))
                        company_data['created_by'] = user
                        company_data['last_modified_by'] = user
                        company_data['status'] = CompanyStatus.COMPLETE
                        
                        def create_company():
                            with transaction.atomic():
                                comp = Company.objects.create(**company_data)
                                CompanyHistory.objects.create(
                                    company=comp,
                                    action='CREATED',
                                    performed_by=user,
                                    description=f'Imported from Excel sheet: {sheet_name}'
                                )
                                return comp
                        
                        company = retry_on_locked(create_company)
                        
                        # Add to caches
                        exact_match_cache[comp_hash] = (company, company_data)
                        if simple_key not in simple_key_cache:
                            simple_key_cache[simple_key] = []
                        simple_key_cache[simple_key].append((company, company_data, category))
                        
                        results['companies_created'] += 1
                    
                    # Check if this category already exists for this company
                    existing_site = ProductionSite.objects.filter(
                        company=company,
                        category=category
                    ).first()
                    
                    if existing_site:
                        # Update existing version with retry
                        current_version = existing_site.versions.filter(is_current=True).first()
                        if current_version:
                            version_data = cls._extract_version_data(row_data, valid_version_fields)
                            for field, value in version_data.items():
                                if value is not None:
                                    setattr(current_version, field, value)
                            
                            def save_version():
                                current_version.save()
                            retry_on_locked(save_version)
                    else:
                        # Create new production site and Initial Version (version_number=0)
                        def create_site_and_version():
                            with transaction.atomic():
                                site = ProductionSite.objects.create(
                                    company=company,
                                    category=category,
                                    created_by=user
                                )
                                
                                version_data = cls._extract_version_data(row_data, valid_version_fields)
                                
                                # CRITICAL: Clean data for PostgreSQL compatibility
                                version_data = clean_version_data_for_postgres(version_data, ProductionSiteVersion)
                                
                                version_data['production_site'] = site
                                version_data['version_number'] = 0  # Initial Version
                                version_data['is_current'] = True
                                version_data['is_active'] = True
                                version_data['is_initial'] = True  # Mark as initial version
                                version_data['created_by'] = user
                                version_data['version_notes'] = 'Initial Version'
                                
                                # Build snapshot data for the Initial Version
                                # Company data snapshot
                                company_snapshot = {}
                                for field in COMMON_FIELDS:
                                    company_snapshot[field] = getattr(company, field, '') or ''
                                version_data['company_data_snapshot'] = company_snapshot
                                
                                # Contact data snapshot
                                contact_snapshot = {}
                                for field in CONTACT_FIELDS:
                                    contact_snapshot[field] = getattr(company, field, '') or ''
                                version_data['contact_data_snapshot'] = contact_snapshot
                                
                                # Notes snapshot (empty for new imports)
                                version_data['notes_snapshot'] = []
                                
                                ProductionSiteVersion.objects.create(**version_data)
                                return site
                        
                        site = retry_on_locked(create_site_and_version)
                        if site:
                            results['sites_created'] += 1
                            results['versions_created'] += 1
                
                except Exception as e:
                    results['errors'].append({
                        'sheet': sheet_name,
                        'row': row_idx,
                        'company': row_data.get('company_name', 'Unknown'),
                        'error': str(e)
                    })
        
        wb.close()
        
        # Generate potential duplicates report if requested and there are duplicates
        if report_path and results['potential_duplicates']:
            cls._generate_duplicates_report(results['potential_duplicates'], report_path)
            results['report_path'] = report_path
        
        return results
    
    @classmethod
    def _get_category_from_sheet_name(cls, sheet_name):
        """Get category code from sheet name."""
        sheet_lower = sheet_name.lower().strip()
        return SHEET_TO_CATEGORY.get(sheet_lower)
    
    @classmethod
    def _extract_company_data(cls, row_data):
        """Extract company-level fields from row data."""
        data = {}
        for field in COMPANY_FIELDS:
            value = row_data.get(field)
            if value is not None and str(value).strip():
                data[field] = str(value).strip()
            else:
                data[field] = ''
        return data
    
    @classmethod
    def _extract_version_data(cls, row_data, valid_fields):
        """Extract version-level fields from row data, only for valid fields.
        CRITICAL: Properly handles all field types for PostgreSQL compatibility."""
        data = {}
        for field, value in row_data.items():
            # Skip company fields
            if field in COMPANY_FIELDS:
                continue
            
            # Skip fields not valid for this category
            if field not in valid_fields:
                continue
            
            # Parse based on field type
            if field in INTEGER_FIELDS:
                parsed = parse_integer(value)
                if parsed is not None:
                    data[field] = parsed
            elif field in TEXT_FIELDS:
                if value is not None and str(value).strip():
                    data[field] = str(value).strip()
            else:
                # Boolean field - ALWAYS convert to True/False
                # This is critical for PostgreSQL
                data[field] = parse_boolean(value)
        
        return data
    
    @classmethod
    def _generate_duplicates_report(cls, potential_duplicates, report_path):
        """Generate Excel report of potential duplicates for manual review."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Potential Duplicates'
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        new_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        existing_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        diff_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            'Duplicate #', 'Field', 
            'NEW - Category', 'NEW - Sheet', 'NEW - Row', 'NEW - Value',
            'EXISTING - Key', 'EXISTING - Categories', 'EXISTING - Value',
            'Action Needed'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Data
        row_idx = 2
        for dup_idx, dup in enumerate(potential_duplicates, 1):
            new_rec = dup['new_record']
            existing_rec = dup['existing_record']
            differences = dup['differences']
            
            # Write company name row
            ws.cell(row=row_idx, column=1, value=dup_idx).border = thin_border
            ws.cell(row=row_idx, column=2, value='company_name').border = thin_border
            ws.cell(row=row_idx, column=3, value=new_rec['category']).border = thin_border
            ws.cell(row=row_idx, column=3).fill = new_fill
            ws.cell(row=row_idx, column=4, value=new_rec['sheet']).border = thin_border
            ws.cell(row=row_idx, column=4).fill = new_fill
            ws.cell(row=row_idx, column=5, value=new_rec['row']).border = thin_border
            ws.cell(row=row_idx, column=5).fill = new_fill
            ws.cell(row=row_idx, column=6, value=new_rec['data'].get('company_name', '')).border = thin_border
            ws.cell(row=row_idx, column=6).fill = new_fill
            ws.cell(row=row_idx, column=7, value=existing_rec['unique_key']).border = thin_border
            ws.cell(row=row_idx, column=7).fill = existing_fill
            ws.cell(row=row_idx, column=8, value=', '.join(existing_rec['categories'])).border = thin_border
            ws.cell(row=row_idx, column=8).fill = existing_fill
            ws.cell(row=row_idx, column=9, value=existing_rec['data'].get('company_name', '')).border = thin_border
            ws.cell(row=row_idx, column=9).fill = existing_fill
            ws.cell(row=row_idx, column=10, value='REVIEW').border = thin_border
            row_idx += 1
            
            # Write difference rows
            for diff in differences:
                ws.cell(row=row_idx, column=1, value='').border = thin_border
                ws.cell(row=row_idx, column=2, value=diff['field']).border = thin_border
                ws.cell(row=row_idx, column=2).fill = diff_fill
                ws.cell(row=row_idx, column=3, value='').border = thin_border
                ws.cell(row=row_idx, column=4, value='').border = thin_border
                ws.cell(row=row_idx, column=5, value='').border = thin_border
                ws.cell(row=row_idx, column=6, value=diff['value1'] or '(empty)').border = thin_border
                ws.cell(row=row_idx, column=6).fill = diff_fill
                ws.cell(row=row_idx, column=7, value='').border = thin_border
                ws.cell(row=row_idx, column=8, value='').border = thin_border
                ws.cell(row=row_idx, column=9, value=diff['value2'] or '(empty)').border = thin_border
                ws.cell(row=row_idx, column=9).fill = diff_fill
                ws.cell(row=row_idx, column=10, value='').border = thin_border
                row_idx += 1
            
            # Add blank row between duplicates
            row_idx += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 40
        ws.column_dimensions['J'].width = 15
        
        # Freeze header
        ws.freeze_panes = 'A2'
        
        # Save
        wb.save(report_path)
        wb.close()
        
        print(f"Generated potential duplicates report: {report_path}")
