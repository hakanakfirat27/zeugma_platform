# reports/import_export_views.py
"""
API Views for bulk import/export operations with real field metadata
FIXED VERSION - Properly handles boolean fields (1/0/blank) and prevents NULL constraint errors
"""

import logging
import pandas as pd
import numpy as np
import io
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Q
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import UnverifiedSite, DataCollectionProject
from .fields import (
    COMMON_FIELDS,
    CONTACT_FIELDS,
    INJECTION_FIELDS,
    BLOW_FIELDS,
    ROTO_FIELDS,
    PE_FILM_FIELDS,
    SHEET_FIELDS,
    PIPE_FIELDS,
    TUBE_HOSE_FIELDS,
    PROFILE_FIELDS,
    CABLE_FIELDS,
    COMPOUNDER_FIELDS,
)
from .field_metadata_view import get_field_metadata

logger = logging.getLogger(__name__)

# Excel styling constants
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Category field mappings
CATEGORY_FIELD_MAP = {
    'INJECTION': INJECTION_FIELDS,
    'BLOW': BLOW_FIELDS,
    'ROTO': ROTO_FIELDS,
    'PE_FILM': PE_FILM_FIELDS,
    'SHEET': SHEET_FIELDS,
    'PIPE': PIPE_FIELDS,
    'TUBE_HOSE': TUBE_HOSE_FIELDS,
    'PROFILE': PROFILE_FIELDS,
    'CABLE': CABLE_FIELDS,
    'COMPOUNDER': COMPOUNDER_FIELDS,
}

# Category display names
CATEGORY_DISPLAY_MAP = {
    'INJECTION': 'Injection Moulders',
    'BLOW': 'Blow Moulders',
    'ROTO': 'Roto Moulders',
    'PE_FILM': 'PE Film Extruders',
    'SHEET': 'Sheet Extruders',
    'PIPE': 'Pipe Extruders',
    'TUBE_HOSE': 'Tube & Hose Extruders',
    'PROFILE': 'Profile Extruders',
    'CABLE': 'Cable Extruders',
    'COMPOUNDER': 'Compounders',
}


def get_field_type(model, field_name):
    """
    Get the type of a field from the model
    Returns: 'boolean', 'text', 'number', etc.
    """
    try:
        field = model._meta.get_field(field_name)
        field_type = field.get_internal_type()
        
        if field_type in ['BooleanField', 'NullBooleanField']:
            return 'boolean'
        elif field_type in ['CharField', 'TextField', 'EmailField', 'URLField']:
            return 'text'
        elif field_type in ['IntegerField', 'FloatField', 'DecimalField', 'PositiveIntegerField']:
            return 'number'
        else:
            return 'other'
    except:
        return 'unknown'


def get_export_fields_for_category(category):
    """
    Get all export fields in order: Common + Contact + Category specific
    Returns list of (field_name, verbose_name) tuples
    """
    model = UnverifiedSite
    fields_list = []
    
    # Add Common Fields
    for field_name in COMMON_FIELDS:
        metadata = get_field_metadata(model, field_name)
        fields_list.append((field_name, metadata['label']))
    
    # Add Contact Fields
    for field_name in CONTACT_FIELDS:
        metadata = get_field_metadata(model, field_name)
        fields_list.append((field_name, metadata['label']))
    
    # Add Category Specific Fields
    category_fields = CATEGORY_FIELD_MAP.get(category, [])
    for field_name in category_fields:
        metadata = get_field_metadata(model, field_name)
        fields_list.append((field_name, metadata['label']))
    
    return fields_list


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_import_template(request):
    """
    Download Excel template for bulk import with category-specific fields
    
    Query params:
    - category: Category code (INJECTION, BLOW, etc.)
    """
    category = request.GET.get('category', 'INJECTION').upper()
    
    if category not in CATEGORY_FIELD_MAP:
        return Response(
            {'error': 'Invalid category'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Get fields for this category
    export_fields = get_export_fields_for_category(category)
    
    # Create workbook
    wb = Workbook()
    
    # ===== TEMPLATE SHEET =====
    ws_template = wb.active
    ws_template.title = CATEGORY_DISPLAY_MAP[category]
    
    # Write headers with verbose names
    for col_idx, (field_name, verbose_name) in enumerate(export_fields, start=1):
        cell = ws_template.cell(row=1, column=col_idx, value=verbose_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    
    # Adjust column widths
    for col_idx in range(1, len(export_fields) + 1):
        column_letter = get_column_letter(col_idx)
        ws_template.column_dimensions[column_letter].width = 20
    
    # Freeze header row
    ws_template.freeze_panes = 'A2'
    
    # ===== INSTRUCTIONS SHEET =====
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions = [
        ("BULK IMPORT INSTRUCTIONS", "", ""),
        ("", "", ""),
        ("How to use this template:", "", ""),
        ("1. Fill in your data starting from row 2", "", ""),
        ("2. Do not modify the header row", "", ""),
        ("3. Required fields must be filled in", "", ""),
        ("4. Upload the completed file for import", "", ""),
        ("", "", ""),
        ("Field Descriptions:", "", ""),
        ("- Company Name: Legal or trading name (REQUIRED)", "", ""),
        ("- Country: Country of operation (REQUIRED)", "", ""),
        ("- Checkbox fields: Use 1 for Yes, 0 for No, leave BLANK if unknown", "", ""),
        ("- Text fields: Free text entry, leave BLANK if unknown", "", ""),
        ("- Number fields: Numeric values only", "", ""),
        ("", "", ""),
        ("Important Notes:", "", ""),
        ("- Duplicates will be SKIPPED (same company name in same project)", "", ""),
        ("- Leave optional fields BLANK if unknown (do not use 0, False, or No)", "", ""),
        ("- For checkboxes: 1=Yes, 0=No, blank=Unknown (will default to No)", "", ""),
        ("", "", ""),
        ("Support:", "", ""),
        ("Contact your administrator for help", "", ""),
    ]
    
    # Write instructions
    for row_idx, (col1, col2, col3) in enumerate(instructions, start=1):
        ws_instructions.cell(row=row_idx, column=1, value=col1)
        ws_instructions.cell(row=row_idx, column=2, value=col2)
        ws_instructions.cell(row=row_idx, column=3, value=col3)
        
        # Style title row
        if row_idx == 1:
            for col in range(1, 4):
                cell = ws_instructions.cell(row=row_idx, column=col)
                cell.font = Font(bold=True, size=14, color="4472C4")
    
    # Adjust column widths for instructions
    ws_instructions.column_dimensions['A'].width = 35
    ws_instructions.column_dimensions['B'].width = 50
    ws_instructions.column_dimensions['C'].width = 40
    
    # Save to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"{CATEGORY_DISPLAY_MAP[category]}_import_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    logger.info(f"Import template for {category} downloaded by {request.user.username}")
    
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_sites(request):
    """
    Export sites to Excel with proper field metadata
    FIXED: Boolean fields export correctly as 1/0/blank
    
    Body:
    {
        "project_id": "uuid", (required)
        "category": "INJECTION", (optional - filter by category)
        "verification_status": ["PENDING"], (optional)
        "format": "xlsx" or "csv" (default: xlsx)
    }
    """
    
    user = request.user
    
    # Get parameters
    project_id = request.data.get('project_id')
    if not project_id:
        return Response(
            {'error': 'Project ID is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    category_filter = request.data.get('category', '').upper()
    verification_statuses = request.data.get('verification_status', [])
    export_format = request.data.get('format', 'xlsx')
    
    # Verify project access
    try:
        project = DataCollectionProject.objects.get(project_id=project_id)
        if user.role == 'DATA_COLLECTOR' and project.created_by != user:
            return Response(
                {'error': 'You do not have access to this project'},
                status=status.HTTP_403_FORBIDDEN
            )
    except DataCollectionProject.DoesNotExist:
        return Response(
            {'error': 'Project not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Build queryset
    queryset = UnverifiedSite.objects.filter(project=project).order_by('category', 'company_name')
    
    # Apply filters
    if category_filter and category_filter in CATEGORY_FIELD_MAP:
        queryset = queryset.filter(category=category_filter)
    
    if verification_statuses:
        queryset = queryset.filter(verification_status__in=verification_statuses)
    
    if queryset.count() == 0:
        return Response(
            {'error': 'No sites found matching the filters'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Group sites by category
    sites_by_category = {}
    for site in queryset:
        if site.category not in sites_by_category:
            sites_by_category[site.category] = []
        sites_by_category[site.category].append(site)
    
    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create a sheet for each category
    for category, sites in sites_by_category.items():
        # Get fields for this category
        export_fields = get_export_fields_for_category(category)
        
        # Create sheet with category display name
        ws = wb.create_sheet(CATEGORY_DISPLAY_MAP.get(category, category))
        
        # Write headers
        for col_idx, (field_name, verbose_name) in enumerate(export_fields, start=1):
            cell = ws.cell(row=1, column=col_idx, value=verbose_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER
        
        # Write data rows
        for row_idx, site in enumerate(sites, start=2):
            for col_idx, (field_name, verbose_name) in enumerate(export_fields, start=1):
                # Get value from site
                value = getattr(site, field_name, None)
                
                # Get field type
                field_type = get_field_type(UnverifiedSite, field_name)
                
                # FIXED: Format values based on type
                if field_type == 'boolean':
                    # Convert boolean to 1/0/blank for Excel
                    if value is True:
                        value = 1  # True = 1
                    elif value is False:
                        value = 0  # False = 0
                    else:  # None or null
                        value = ''  # Empty = blank (not 0!)
                elif value is None or value == '':
                    # Empty text/number fields should be blank
                    value = ''
                else:
                    # Convert other values to string
                    value = str(value)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = BORDER
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Adjust column widths
        for col_idx in range(1, len(export_fields) + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 20
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    # Save to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create filename from project name
    filename = f"{project.project_name}.xlsx"
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    logger.info(
        f"Exported {queryset.count()} sites from project {project_id} "
        f"by {user.username}"
    )
    
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_preview(request):
    """
    Preview and validate import data before confirming
    FIXED: Properly handles boolean fields - converts blanks to False (not None)
    
    Expects multipart/form-data with:
    - file: Excel/CSV file
    - project_id: UUID of project
    - category: Category code (INJECTION, BLOW, etc.)
    """
    
    user = request.user
    project_id = request.data.get('project_id')
    category = request.data.get('category', '').upper()
    uploaded_file = request.FILES.get('file')
    
    # Validation
    if not project_id:
        return Response(
            {'error': 'Project ID is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    if not category or category not in CATEGORY_FIELD_MAP:
        return Response(
            {'error': 'Valid category is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    if not uploaded_file:
        return Response(
            {'error': 'No file uploaded'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Verify project
    try:
        project = DataCollectionProject.objects.get(project_id=project_id)
        if user.role == 'DATA_COLLECTOR' and project.created_by != user:
            return Response(
                {'error': 'You do not have access to this project'},
                status=status.HTTP_403_FORBIDDEN
            )
    except DataCollectionProject.DoesNotExist:
        return Response(
            {'error': 'Project not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Read file
    try:
        file_ext = uploaded_file.name.split('.')[-1].lower()
        if file_ext in ['xlsx', 'xls']:
            df = pd.read_excel(uploaded_file)
        elif file_ext == 'csv':
            df = pd.read_csv(uploaded_file)
        else:
            return Response(
                {'error': 'Invalid file format. Please upload .xlsx, .xls, or .csv file'},
                status=status.HTTP_400_BAD_REQUEST
            )
    except Exception as e:
        logger.error(f"Error reading file: {str(e)}")
        return Response(
            {'error': f'Error reading file: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Get expected fields for this category
    export_fields = get_export_fields_for_category(category)
    field_name_to_verbose = {field_name: verbose_name for field_name, verbose_name in export_fields}
    verbose_to_field_name = {verbose_name: field_name for field_name, verbose_name in export_fields}
    
    # Map column headers to field names
    df_columns_mapped = {}
    for col in df.columns:
        col_stripped = str(col).strip()
        if col_stripped in verbose_to_field_name:
            df_columns_mapped[col] = verbose_to_field_name[col_stripped]
    
    # Rename columns
    df.rename(columns=df_columns_mapped, inplace=True)
    
    # Replace NaN with None
    df = df.replace({np.nan: None})
    
    # Validate rows
    errors = []
    valid_rows = []
    warnings = []
    duplicate_rows = []
    
    for idx, row in df.iterrows():
        row_number = idx + 2  # Excel row number
        row_data = row.to_dict()
        
        # Check required fields
        company_name = str(row_data.get('company_name', '')).strip() if row_data.get('company_name') is not None else ''
        country = str(row_data.get('country', '')).strip() if row_data.get('country') is not None else ''
        
        if not company_name:
            errors.append({
                'row': row_number,
                'field': 'company_name',
                'error': 'Company name is required'
            })
            continue
        
        if not country:
            errors.append({
                'row': row_number,
                'field': 'country',
                'error': 'Country is required'
            })
            continue
        
        # Check for duplicates
        duplicate_count = UnverifiedSite.objects.filter(
            company_name__iexact=company_name,
            project=project
        ).count()
        
        if duplicate_count > 0:
            duplicate_rows.append(row_number)
            warnings.append({
                'row': row_number,
                'message': f'Duplicate: "{company_name}" already exists {duplicate_count} time(s) in this project - WILL BE SKIPPED'
            })
            continue
        
        # FIXED: Process each field based on its type
        processed_data = {}
        for field_name, value in row_data.items():
            if field_name not in [f[0] for f in export_fields]:
                continue  # Skip unknown fields
            
            field_type = get_field_type(UnverifiedSite, field_name)
            
            if field_type == 'boolean':
                # CRITICAL FIX: Convert blanks to False, not None
                # This prevents the NOT NULL constraint error
                if value is None or value == '':
                    processed_data[field_name] = False  # Default to False for blank values
                elif isinstance(value, str):
                    value_stripped = value.strip().lower()
                    if value_stripped == 'nan' or value_stripped == '':
                        processed_data[field_name] = False  # Default to False
                    elif value_stripped in ['yes', 'true', '1', 'y']:
                        processed_data[field_name] = True
                    elif value_stripped in ['no', 'false', '0', 'n']:
                        processed_data[field_name] = False
                    else:
                        processed_data[field_name] = False  # Default to False for unknown values
                elif isinstance(value, (int, float)):
                    # Excel stores 1/0 as numbers
                    if value == 1:
                        processed_data[field_name] = True
                    elif value == 0:
                        processed_data[field_name] = False
                    else:
                        processed_data[field_name] = False  # Default to False
                elif isinstance(value, bool):
                    processed_data[field_name] = value
                else:
                    processed_data[field_name] = False  # Default to False
            
            elif field_type == 'text':
                # Handle text fields
                if value is None:
                    processed_data[field_name] = ''
                elif isinstance(value, str):
                    value_stripped = value.strip()
                    if value_stripped.lower() == 'nan' or value_stripped == '':
                        processed_data[field_name] = ''
                    else:
                        processed_data[field_name] = value_stripped
                else:
                    # Convert numbers/other types to string
                    processed_data[field_name] = str(value).strip()
            
            elif field_type == 'number':
                # Handle numeric fields
                if value is None or value == '':
                    processed_data[field_name] = None
                else:
                    try:
                        if isinstance(value, str) and value.strip().lower() == 'nan':
                            processed_data[field_name] = None
                        else:
                            processed_data[field_name] = float(value) if '.' in str(value) else int(value)
                    except:
                        processed_data[field_name] = None
            
            else:
                # Other fields - keep as is, but handle "nan"
                if value is None:
                    processed_data[field_name] = ''
                elif isinstance(value, str) and value.strip().lower() == 'nan':
                    processed_data[field_name] = ''
                else:
                    processed_data[field_name] = value
        
        valid_rows.append({
            'row_number': row_number,
            'data': processed_data
        })
    
    # Prepare preview
    preview_data = []
    for row in valid_rows[:10]:
        preview_data.append({
            'company_name': row['data'].get('company_name', ''),
            'country': row['data'].get('country', ''),
            'address_1': row['data'].get('address_1', ''),
        })
    
    # Store validated data in session
    request.session[f'import_data_{project_id}'] = {
        'valid_rows': valid_rows,
        'category': category,
        'duplicate_rows': duplicate_rows,
        'timestamp': datetime.now().isoformat(),
    }
    
    response_data = {
        'total_rows': len(df),
        'valid_rows': len(valid_rows),
        'invalid_rows': len(errors),
        'duplicate_rows': len(duplicate_rows),
        'errors': errors[:50],
        'preview_data': preview_data,
        'warnings': warnings[:50],
        'has_more_errors': len(errors) > 50,
        'has_more_warnings': len(warnings) > 50,
    }
    
    logger.info(
        f"Import preview: {len(valid_rows)} valid, {len(errors)} invalid, {len(duplicate_rows)} duplicate rows "
        f"for project {project_id} category {category} by {user.username}"
    )
    
    return Response(response_data, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_confirm(request):
    """
    Confirm and execute the import after preview
    FIXED: Creates sites with proper boolean values (no NULLs)
    
    Body:
    {
        "project_id": "uuid"
    }
    """
    
    user = request.user
    project_id = request.data.get('project_id')
    
    if not project_id:
        return Response(
            {'error': 'Project ID is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Get validated data from session
    import_data = request.session.get(f'import_data_{project_id}')
    
    if not import_data:
        return Response(
            {'error': 'No import data found. Please run preview first.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Check if data is fresh (within 1 hour)
    timestamp = datetime.fromisoformat(import_data['timestamp'])
    if (datetime.now() - timestamp).total_seconds() > 3600:
        return Response(
            {'error': 'Import data expired. Please run preview again.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Verify project
    try:
        project = DataCollectionProject.objects.get(project_id=project_id)
        if user.role == 'DATA_COLLECTOR' and project.created_by != user:
            return Response(
                {'error': 'You do not have access to this project'},
                status=status.HTTP_403_FORBIDDEN
            )
    except DataCollectionProject.DoesNotExist:
        return Response(
            {'error': 'Project not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Import sites
    valid_rows = import_data['valid_rows']
    category = import_data['category']
    duplicate_rows = import_data.get('duplicate_rows', [])
    imported_sites = []
    skipped_duplicates = 0
    
    try:
        for row in valid_rows:
            site_data = row['data']
            company_name = site_data.get('company_name', '')
            
            # Double-check for duplicates before creating
            if UnverifiedSite.objects.filter(
                company_name__iexact=company_name,
                project=project
            ).exists():
                skipped_duplicates += 1
                continue
            
            # Create site instance
            site = UnverifiedSite(
                project=project,
                collected_by=user,
                category=category,
                verification_status='PENDING',
            )
            
            # Set all fields from data
            for field_name, value in site_data.items():
                if hasattr(site, field_name):
                    # IMPORTANT: The value has already been processed in preview
                    # Boolean fields are True/False (never None)
                    # Text fields are strings (empty string if blank)
                    # Number fields can be None
                    setattr(site, field_name, value)
            
            site.save()
            imported_sites.append(site)
        
        # Clear session data
        del request.session[f'import_data_{project_id}']
        
        message = f'{len(imported_sites)} sites imported successfully'
        if skipped_duplicates > 0:
            message += f' ({skipped_duplicates} duplicates skipped)'
        
        logger.info(
            f"{len(imported_sites)} sites imported ({skipped_duplicates} skipped) to project {project_id} "
            f"category {category} by {user.username}"
        )
        
        return Response({
            'success': True,
            'imported_count': len(imported_sites),
            'skipped_count': skipped_duplicates,
            'message': message
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error during import: {str(e)}")
        return Response(
            {'error': f'Error during import: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
