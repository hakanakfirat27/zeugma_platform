"""
API Views for Export Template functionality.
"""

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import ExportTemplate, CustomReport, Subscription, SuperdatabaseRecord
from .serializers import ExportTemplateSerializer, ExportTemplateCreateSerializer
from .filters import SuperdatabaseRecordFilter
from accounts.models import UserRole
from django.db.models import Q


class ExportTemplateListCreateAPIView(APIView):
    """List and create export templates"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get all export templates for a specific report"""

        if request.user.role != UserRole.CLIENT:
            return Response(
                {"error": "Only clients can access export templates"},
                status=status.HTTP_403_FORBIDDEN
            )

        report_id = request.query_params.get('report_id')

        if not report_id:
            return Response(
                {"error": "report_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verify report exists
        try:
            report = CustomReport.objects.get(report_id=report_id)
        except CustomReport.DoesNotExist:
            return Response(
                {"error": "Report not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Verify access
        today = timezone.now().date()
        has_subscription = Subscription.objects.filter(
            client=request.user,
            report=report,
            status='ACTIVE',
            start_date__lte=today,
            end_date__gte=today
        ).exists()

        if not has_subscription:
            return Response(
                {"error": "You don't have access to this report"},
                status=status.HTTP_403_FORBIDDEN
            )

        # Get templates
        templates = ExportTemplate.objects.filter(
            user=request.user,
            report=report
        )

        serializer = ExportTemplateSerializer(templates, many=True)

        return Response({
            "count": templates.count(),
            "results": serializer.data
        })

    def post(self, request):
        """Create a new export template"""

        if request.user.role != UserRole.CLIENT:
            return Response(
                {"error": "Only clients can create export templates"},
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = ExportTemplateCreateSerializer(data=request.data)

        if not serializer.is_valid():
            return Response(
                {"error": "Invalid data", "details": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )

        validated_data = serializer.validated_data
        report = get_object_or_404(CustomReport, report_id=validated_data['report_id'])

        # Verify access
        today = timezone.now().date()
        has_subscription = Subscription.objects.filter(
            client=request.user,
            report=report,
            status='ACTIVE',
            start_date__lte=today,
            end_date__gte=today
        ).exists()

        if not has_subscription:
            return Response(
                {"error": "You don't have access to this report"},
                status=status.HTTP_403_FORBIDDEN
            )

        # Check for duplicate name
        if ExportTemplate.objects.filter(
                user=request.user,
                report=report,
                name=validated_data['name']
        ).exists():
            return Response(
                {"error": f"A template named '{validated_data['name']}' already exists"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # If setting as default, unset other defaults
        if validated_data.get('is_default', False):
            ExportTemplate.objects.filter(
                user=request.user,
                report=report,
                is_default=True
            ).update(is_default=False)

        # Create template
        template = ExportTemplate.objects.create(
            user=request.user,
            report=report,
            name=validated_data['name'],
            description=validated_data.get('description', ''),
            selected_columns=validated_data['selected_columns'],
            is_default=validated_data.get('is_default', False)
        )

        response_serializer = ExportTemplateSerializer(template)

        return Response({
            "message": "Export template created successfully",
            "template": response_serializer.data
        }, status=status.HTTP_201_CREATED)


class ExportTemplateDetailAPIView(APIView):
    """Retrieve, update, and delete export templates"""

    permission_classes = [IsAuthenticated]

    def get_object(self, template_id):
        """Get template and verify ownership"""
        try:
            return ExportTemplate.objects.get(
                id=template_id,
                user=self.request.user
            )
        except ExportTemplate.DoesNotExist:
            return None

    def get(self, request, template_id):
        """Get template details"""

        if request.user.role != UserRole.CLIENT:
            return Response(
                {"error": "Only clients can access export templates"},
                status=status.HTTP_403_FORBIDDEN
            )

        template = self.get_object(template_id)

        if not template:
            return Response(
                {"error": "Template not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = ExportTemplateSerializer(template)
        return Response(serializer.data)

    def patch(self, request, template_id):
        """Update template"""

        if request.user.role != UserRole.CLIENT:
            return Response(
                {"error": "Only clients can update export templates"},
                status=status.HTTP_403_FORBIDDEN
            )

        template = self.get_object(template_id)

        if not template:
            return Response(
                {"error": "Template not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Update fields
        if 'name' in request.data:
            new_name = request.data['name'].strip()
            if not new_name:
                return Response(
                    {"error": "Name cannot be empty"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Check duplicate
            if ExportTemplate.objects.filter(
                    user=request.user,
                    report=template.report,
                    name=new_name
            ).exclude(id=template_id).exists():
                return Response(
                    {"error": f"A template named '{new_name}' already exists"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            template.name = new_name

        if 'description' in request.data:
            template.description = request.data['description']

        if 'selected_columns' in request.data:
            columns = request.data['selected_columns']
            if not isinstance(columns, list) or len(columns) == 0:
                return Response(
                    {"error": "Please select at least one column"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            template.selected_columns = columns

        if 'is_default' in request.data:
            is_default = request.data['is_default']
            if is_default:
                ExportTemplate.objects.filter(
                    user=request.user,
                    report=template.report,
                    is_default=True
                ).exclude(id=template_id).update(is_default=False)
            template.is_default = is_default

        template.save()

        serializer = ExportTemplateSerializer(template)

        return Response({
            "message": "Template updated successfully",
            "template": serializer.data
        })

    def delete(self, request, template_id):
        """Delete template"""

        if request.user.role != UserRole.CLIENT:
            return Response(
                {"error": "Only clients can delete export templates"},
                status=status.HTTP_403_FORBIDDEN
            )

        template = self.get_object(template_id)

        if not template:
            return Response(
                {"error": "Template not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        template_name = template.name
        template.delete()

        return Response({
            "message": f"Template '{template_name}' deleted successfully"
        })


class ExportToExcelAPIView(APIView):
    """Export data to Excel with selected columns"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Export data to Excel"""

        if request.user.role != UserRole.CLIENT:
            return Response(
                {"error": "Only clients can export data"},
                status=status.HTTP_403_FORBIDDEN
            )

        report_id = request.query_params.get('report_id')
        template_id = request.query_params.get('template_id')

        if not report_id:
            return Response(
                {"error": "report_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get report and verify access
        try:
            report = CustomReport.objects.get(report_id=report_id)
        except CustomReport.DoesNotExist:
            return Response(
                {"error": "Report not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        today = timezone.now().date()
        subscription = Subscription.objects.filter(
            client=request.user,
            report=report,
            status='ACTIVE',
            start_date__lte=today,
            end_date__gte=today
        ).first()

        if not subscription:
            return Response(
                {"error": "No active subscription"},
                status=status.HTTP_403_FORBIDDEN
            )

        # Get selected columns
        if template_id:
            try:
                template = ExportTemplate.objects.get(
                    id=template_id,
                    user=request.user
                )
                selected_columns = template.selected_columns
            except ExportTemplate.DoesNotExist:
                return Response(
                    {"error": "Template not found"},
                    status=status.HTTP_404_NOT_FOUND
                )
        else:
            # Get columns from query params
            columns_param = request.query_params.get('columns')
            if columns_param:
                selected_columns = columns_param.split(',')
            else:
                # Default columns
                selected_columns = [
                    'company_name', 'country', 'address_1',
                    'email', 'telephone', 'category'
                ]

        # Get filtered records
        filter_criteria = report.filter_criteria or {}
        queryset = SuperdatabaseRecord.objects.all()

        # Apply report filters
        filter_q = Q()
        for field, value in filter_criteria.items():
            if field == 'categories':
                if isinstance(value, list) and len(value) > 0:
                    filter_q &= Q(category__in=value)
                elif isinstance(value, str) and value:
                    filter_q &= Q(category=value)
                continue

            if value is not None:
                if isinstance(value, list):
                    if len(value) > 0:
                        filter_q &= Q(**{f"{field}__in": value})
                else:
                    filter_q &= Q(**{field: value})

        if filter_q:
            queryset = queryset.filter(filter_q)

        # Apply user filters
        categories_param = request.query_params.get('categories')
        if categories_param:
            category_list = [c.strip() for c in categories_param.split(',') if c.strip()]
            if category_list:
                queryset = queryset.filter(category__in=category_list)

        filterset = SuperdatabaseRecordFilter(request.GET, queryset=queryset)
        queryset = filterset.qs

        # Limit
        max_records = 10000
        if queryset.count() > max_records:
            return Response(
                {"error": f"Export limited to {max_records} records. Please apply more filters."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report.title[:31]  # Excel sheet name limit

        # Get field labels
        field_labels = {
            field.name: field.verbose_name
            for field in SuperdatabaseRecord._meta.fields
        }

        # Write headers
        headers = [field_labels.get(col, col) for col in selected_columns]

        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Add comment to boolean fields to explain 1=True, 0=False
            field_name = selected_columns[col_num - 1]
            if field_labels.get(field_name) == 'BooleanField':
                from openpyxl.comments import Comment
                cell.comment = Comment("1 = Yes/True\n0 = No/False", "System")

        # Write data
        for row_num, record in enumerate(queryset, 2):
            for col_num, field_name in enumerate(selected_columns, 1):
                value = getattr(record, field_name, '')

                # IMPORTANT: Convert boolean values to 1/0
                if isinstance(value, bool):
                    value = 1 if value else 0
                # Handle None values
                elif value is None:
                    value = ''
                # Convert to string for other types
                else:
                    value = str(value)

                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for col_num in range(1, len(selected_columns) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        filename = f"{report.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)

        return response