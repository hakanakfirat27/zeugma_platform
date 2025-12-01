# reports/management/commands/list_merged_companies.py
"""
List all companies with multiple production categories.
"""

from django.core.management.base import BaseCommand
from reports.company_models import Company


class Command(BaseCommand):
    help = 'List all companies with multiple production categories'

    def add_arguments(self, parser):
        parser.add_argument(
            '--min-categories',
            type=int,
            default=2,
            help='Minimum number of categories (default: 2)'
        )
        parser.add_argument(
            '--country',
            type=str,
            help='Filter by country'
        )
        parser.add_argument(
            '--export',
            type=str,
            help='Export to CSV file path'
        )

    def handle(self, *args, **options):
        min_cats = options['min_categories']
        country_filter = options.get('country')
        export_path = options.get('export')
        
        # Get all companies with their production sites
        companies = Company.objects.prefetch_related('production_sites').all()
        
        if country_filter:
            companies = companies.filter(country__iexact=country_filter)
        
        # Find companies with multiple categories
        merged_companies = []
        for company in companies:
            categories = list(company.production_sites.values_list('category', flat=True))
            if len(categories) >= min_cats:
                merged_companies.append({
                    'unique_key': company.unique_key,
                    'company_name': company.company_name,
                    'country': company.country,
                    'address_1': company.address_1,
                    'num_categories': len(categories),
                    'categories': sorted(categories)
                })
        
        # Sort by number of categories (descending), then by name
        merged_companies.sort(key=lambda x: (-x['num_categories'], x['company_name']))
        
        # Display results
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(f'Companies with {min_cats}+ categories: {len(merged_companies)}'))
        self.stdout.write('=' * 100)
        
        # Group by number of categories for summary
        by_count = {}
        for mc in merged_companies:
            count = mc['num_categories']
            if count not in by_count:
                by_count[count] = 0
            by_count[count] += 1
        
        self.stdout.write('')
        self.stdout.write('SUMMARY BY CATEGORY COUNT:')
        for count in sorted(by_count.keys(), reverse=True):
            self.stdout.write(f'  {count} categories: {by_count[count]} companies')
        
        self.stdout.write('')
        self.stdout.write('=' * 100)
        self.stdout.write('')
        
        # Print all merged companies
        for mc in merged_companies:
            cats_str = ', '.join(mc['categories'])
            self.stdout.write(
                f"{mc['unique_key']} | {mc['company_name'][:50]:<50} | "
                f"{mc['country']:<20} | {mc['num_categories']} cats | [{cats_str}]"
            )
        
        # Export to Excel if requested
        if export_path:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Merged Companies'
            
            # Define styles
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = ['Unique Key', 'Company Name', 'Country', 'Address', 'Num Categories', 'Categories']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Data rows
            for row_idx, mc in enumerate(merged_companies, 2):
                ws.cell(row=row_idx, column=1, value=mc['unique_key']).border = thin_border
                ws.cell(row=row_idx, column=2, value=mc['company_name']).border = thin_border
                ws.cell(row=row_idx, column=3, value=mc['country']).border = thin_border
                ws.cell(row=row_idx, column=4, value=mc['address_1']).border = thin_border
                ws.cell(row=row_idx, column=5, value=mc['num_categories']).border = thin_border
                ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal='center')
                ws.cell(row=row_idx, column=6, value=', '.join(mc['categories'])).border = thin_border
            
            # Set column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 60
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Auto-filter
            ws.auto_filter.ref = f'A1:F{len(merged_companies) + 1}'
            
            # Save
            wb.save(export_path)
            wb.close()
            
            self.stdout.write('')
            self.stdout.write(self.style.SUCCESS(f'Exported {len(merged_companies)} companies to: {export_path}'))
