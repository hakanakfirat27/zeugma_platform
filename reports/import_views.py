# reports/import_views.py
"""
API Views for Company Import functionality.

Endpoints:
- POST /api/companies/import/ - Import companies from Excel
- POST /api/companies/import/validate/ - Validate file before import
- GET /api/companies/import/report/<filename>/ - Download duplicates report
"""

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import FileResponse
from django.utils import timezone
from django.conf import settings
import os
import tempfile


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_companies(request):
    """
    Import companies from Excel file.
    
    The Excel file should have one or more sheets, each corresponding to a category:
    - Sheet names should match category names (e.g., "Injection Moulders", "Blow Moulders")
    - Each sheet must have the 29 common/contact fields as columns
    - Category-specific fields will also be imported
    
    Import Logic:
    1. If a company with ALL 29 common/contact fields matching exists:
       - Only add new production site if the category doesn't exist for that company
       - Skip if category already exists
    2. If company doesn't exist (or fields don't match exactly):
       - Create new company with all data
    
    Returns:
    - Summary of import results
    - Link to download potential duplicates report (if any)
    """
    from .services.company_import import CompanyImportService
    
    # Check if file was uploaded
    if 'file' not in request.FILES:
        return Response(
            {'error': 'No file uploaded. Please upload an Excel file.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    uploaded_file = request.FILES['file']
    
    # Validate file extension
    file_name = uploaded_file.name.lower()
    if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
        return Response(
            {'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls).'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Create temp directory for reports
    import_reports_dir = os.path.join(settings.MEDIA_ROOT, 'import_reports')
    os.makedirs(import_reports_dir, exist_ok=True)
    
    # Generate unique filename for potential duplicates report
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    report_filename = f'potential_duplicates_{timestamp}.xlsx'
    report_path = os.path.join(import_reports_dir, report_filename)
    
    # Save uploaded file to temp location
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        for chunk in uploaded_file.chunks():
            temp_file.write(chunk)
        temp_file_path = temp_file.name
    
    try:
        # Run import
        result = CompanyImportService.import_from_excel(
            file_path=temp_file_path,
            user=request.user,
            report_path=report_path
        )
        
        # Build response
        response_data = {
            'success': True,
            'summary': {
                'total_rows_processed': result.get('total_rows', 0),
                'companies_created': result.get('companies_created', 0),
                'companies_updated': result.get('companies_updated', 0),
                'exact_matches_merged': result.get('merged_count', 0),
                'production_sites_created': result.get('sites_created', 0),
                'versions_created': result.get('versions_created', 0),
                'potential_duplicates_count': len(result.get('potential_duplicates', [])),
                'errors_count': len(result.get('errors', [])),
            },
            'sheets_processed': result.get('sheets_processed', []),
            'errors': result.get('errors', [])[:50],  # Limit errors to 50
        }
        
        # Add report URL if there are potential duplicates
        if result.get('potential_duplicates') and os.path.exists(report_path):
            response_data['potential_duplicates_report'] = f'/api/companies/import/report/{report_filename}'
        
        return Response(response_data, status=status.HTTP_200_OK)
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response(
            {'error': f'Import failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
    finally:
        # Clean up temp file
        if os.path.exists(temp_file_path):
            os.unlink(temp_file_path)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_import_report(request, filename):
    """
    Download the potential duplicates report generated during import.
    """
    import_reports_dir = os.path.join(settings.MEDIA_ROOT, 'import_reports')
    file_path = os.path.join(import_reports_dir, filename)
    
    # Security check - ensure filename doesn't contain path traversal
    if '..' in filename or '/' in filename or '\\' in filename:
        return Response(
            {'error': 'Invalid filename'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    if not os.path.exists(file_path):
        return Response(
            {'error': 'Report not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    return FileResponse(
        open(file_path, 'rb'),
        as_attachment=True,
        filename=filename
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def validate_import_file(request):
    """
    Validate an Excel file before importing.
    
    Returns information about:
    - Detected sheets and their categories
    - Total rows per sheet
    - Any validation errors (missing required columns, etc.)
    """
    import openpyxl
    from .services.company_import import SHEET_TO_CATEGORY, COMPANY_FIELDS, CATEGORY_FIELDS, normalize_header
    
    if 'file' not in request.FILES:
        return Response(
            {'error': 'No file uploaded.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    uploaded_file = request.FILES['file']
    
    # Validate file extension
    file_name = uploaded_file.name.lower()
    if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
        return Response(
            {'error': 'Invalid file type. Please upload an Excel file.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        for chunk in uploaded_file.chunks():
            temp_file.write(chunk)
        temp_file_path = temp_file.name
    
    try:
        wb = openpyxl.load_workbook(temp_file_path, data_only=True, read_only=True)
        
        sheets_info = []
        total_rows = 0
        validation_errors = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Get category from sheet name
            category = SHEET_TO_CATEGORY.get(sheet_name.lower().strip())
            
            # Count rows (excluding header)
            row_count = 0
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                # Check if first cell has value (company name)
                if row[0].value:
                    row_count += 1
            
            total_rows += row_count
            
            # Get headers
            headers = []
            normalized_headers = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append(cell_value)
                    normalized_headers.append(normalize_header(cell_value))
            
            # Check for required company fields
            missing_company_fields = []
            for field in ['company_name', 'country']:  # Minimum required
                if field not in normalized_headers:
                    missing_company_fields.append(field)
            
            sheet_info = {
                'name': sheet_name,
                'category': category,
                'category_recognized': category is not None,
                'row_count': row_count,
                'column_count': len(headers),
                'missing_required_fields': missing_company_fields,
            }
            
            if category and category in CATEGORY_FIELDS:
                # Count how many category-specific fields are present
                category_fields = set(CATEGORY_FIELDS[category])
                present_category_fields = category_fields.intersection(set(normalized_headers))
                sheet_info['category_fields_found'] = len(present_category_fields)
                sheet_info['category_fields_total'] = len(category_fields)
            
            sheets_info.append(sheet_info)
            
            if not category:
                validation_errors.append({
                    'sheet': sheet_name,
                    'error': f'Unknown category. Sheet name "{sheet_name}" does not match any known category.'
                })
            
            if missing_company_fields:
                validation_errors.append({
                    'sheet': sheet_name,
                    'error': f'Missing required fields: {", ".join(missing_company_fields)}'
                })
        
        wb.close()
        
        return Response({
            'valid': len(validation_errors) == 0,
            'file_name': uploaded_file.name,
            'total_sheets': len(sheets_info),
            'total_rows': total_rows,
            'sheets': sheets_info,
            'validation_errors': validation_errors,
        })
    
    except Exception as e:
        return Response(
            {'error': f'Failed to read file: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    finally:
        if os.path.exists(temp_file_path):
            os.unlink(temp_file_path)
